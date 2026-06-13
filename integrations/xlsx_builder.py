"""Генерация Excel (.xlsx) для таблиц данных."""

from __future__ import annotations

import io
import os
from typing import Optional


def build_xlsx_bytes(task_text: str, rows: Optional[list[list]] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    title = (task_text or "Таблица")[:80]
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    headers = ["#", "Название", "Значение", "Статус"]
    header_fill = PatternFill(start_color="4C98FD", end_color="4C98FD", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if not rows:
        rows = [
            [1, "Пункт A", 100, "OK"],
            [2, "Пункт B", 250, "OK"],
            [3, "Пункт C", 75, "Pending"],
            [4, "Пункт D", 180, "OK"],
            [5, "Пункт E", 320, "Review"],
        ]

    for r_idx, row in enumerate(rows, 4):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_xlsx(task_text: str, out_dir: str, filename: str = "table.xlsx", rows=None) -> str:
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, filename)
    with open(path, "wb") as f:
        f.write(build_xlsx_bytes(task_text, rows))
    return path
